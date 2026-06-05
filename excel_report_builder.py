"""Utilidades para construir reportes XLSX con formato a partir de la respuesta Wolkvox."""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe_filename(text: str) -> str:
    text = (text or "").strip() or "server"
    return re.sub(r"[^0-9A-Za-z._-]", "_", text)[:120]


def build_wolkvox_excel(
    *,
    rows: list[Any],
    filename: str,
    sheet_name: str = "report",
) -> tuple[io.BytesIO, str]:
    """Construye un XLSX con cabecera estilo + freeze panes + filtros + anchos.

    - rows puede ser lista de dict o cualquier otro valor.
    - si hay dict, se usan keys como columnas.
    """
    all_keys: list[str] = []
    for r in rows:
        if isinstance(r, dict):
            for k in r.keys():
                if k not in all_keys:
                    all_keys.append(k)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(all_keys or ["raw"])

    for r in rows:
        if isinstance(r, dict):
            ws.append([r.get(k, "") for k in all_keys])
        else:
            ws.append([str(r)])

    # Formato
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    header_row_idx = 1
    col_count = len(all_keys) if all_keys else 1

    ws.freeze_panes = "A2" if ws.max_row >= 2 else "A1"
    ws.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(col_count)}{header_row_idx}"

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    max_width = 45
    for col_idx in range(1, col_count + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(max_width, max(10, max_len + 2))
        ws.column_dimensions[letter].width = width
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = left

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, filename

