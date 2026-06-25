#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script para descargar reportes AMD (campaign_7) para un rango de fechas.
Obtiene los nombres de campaña consultando real_time.php?api=campaigns&campaign_id={ID}
"""

import os
import json
import requests
import io
import re
import time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ========== CONFIGURACIÓN ==========
FECHA_INICIO = "2026-06-20"
FECHA_FIN = datetime.now().strftime("%Y-%m-%d")  # Hoy

SERVIDORES = [
    "operacion-interna",
    "qnt_juridico_blaster",
    "qnt_cobro_blaster",
    "Qnt_RBK_blaster",
    "Qnt_recaudo_blaster",
    "qnt_digital"
]

PAUSA_ENTRE_PETICIONES = 0.5  # Segundos entre consultas de campaña

# ========== CARGA CONFIG ==========
with open('config.json', 'r', encoding='utf-8') as f:
    CONFIG = json.load(f)

BASE_DIR = CONFIG.get('base_dir', r"G:\Unidades compartidas\Analitica\Embudo de Conversión\Proyecto Robot Omnicanal\Producción\2026-06\Resultados")

# ========== LOGGING ==========
import logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler("download_campaign_detail.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ========== FUNCIONES ==========

def get_server(name):
    for s in CONFIG.get('servers', []):
        if s.get('name') == name:
            url = s.get('url', '').strip().rstrip('/')
            token = s.get('token') or CONFIG.get('wolkvox-token')
            if url and token:
                return {'name': name, 'url': url, 'token': token}
    return None

def obtener_nombre_campana_especifica(servidor, campaign_id):
    """Obtiene el nombre de una campaña específica usando campaign_id."""
    try:
        camp_id_clean = re.sub(r'[^0-9]', '', str(campaign_id))
        if not camp_id_clean:
            return None
        
        url = f"{servidor['url']}/api/v2/real_time.php?api=campaigns&campaign_id={campaign_id}"
        headers = {"wolkvox-token": servidor['token']}
        
        resp = requests.get(url, headers=headers, timeout=10)
        
        if resp.status_code == 200:
            data = resp.json()
            campaigns = data.get('data', [])
            if campaigns:
                camp_str = campaigns[0].get('campaign', '')
                if camp_str:
                    parts = camp_str.split(' - ', 1)
                    nombre = parts[1] if len(parts) > 1 else camp_str
                    return nombre
        return None
    except Exception as e:
        logger.debug(f"Error obteniendo nombre para {campaign_id}: {e}")
        return None

def to_ts(fecha, is_end=False):
    try:
        d = datetime.strptime(fecha, '%Y-%m-%d')
        dt = datetime(d.year, d.month, d.day, 23, 59, 59) if is_end else datetime(d.year, d.month, d.day, 0, 0, 0)
        return dt.strftime('%Y%m%d%H%M%S')
    except:
        return fecha

def descargar_reporte(servidor, fecha):
    url = f"{servidor['url']}/api/v2/reports_manager.php?api=campaign_7&campaign_id=all&date_ini={to_ts(fecha)}&date_end={to_ts(fecha, True)}"
    try:
        resp = requests.get(url, headers={"wolkvox-token": servidor['token']}, timeout=60)
        if resp.status_code == 200:
            data = resp.json()
            if data.get('data'):
                return resp.content
        return None
    except Exception as e:
        logger.error(f"Error descargando: {e}")
        return None

def crear_excel_con_nombres(data_bytes, servidor, fecha):
    data = json.loads(data_bytes)
    rows = data.get('data', [])
    if not rows:
        return None
    
    # Recolectar IDs únicos para consultar
    ids_unicos = set()
    for row in rows:
        camp_id = str(row.get('campaign_id', '')).strip()
        if camp_id:
            ids_unicos.add(camp_id)
    
    logger.info(f"   📋 {len(ids_unicos)} IDs de campaña únicos encontrados")
    
    # Obtener nombres para cada ID único
    mapeo = {}
    for camp_id in ids_unicos:
        nombre = obtener_nombre_campana_especifica(servidor, camp_id)
        if nombre:
            camp_id_clean = re.sub(r'[^0-9]', '', camp_id)
            mapeo[camp_id_clean] = nombre
        time.sleep(PAUSA_ENTRE_PETICIONES)
    
    logger.info(f"   📋 {len(mapeo)} nombres obtenidos")
    
    # Reemplazar campaign_id por "ID - Nombre"
    for row in rows:
        camp_id_original = str(row.get('campaign_id', '')).strip()
        camp_id_clean = re.sub(r'[^0-9]', '', camp_id_original)
        
        if camp_id_clean and camp_id_clean in mapeo:
            row['campaign_id'] = f"{camp_id_clean} - {mapeo[camp_id_clean]}"
            row['campaign_name'] = mapeo[camp_id_clean]
        else:
            row['campaign_name'] = ''
    
    # Columnas
    all_keys = set()
    for row in rows:
        all_keys.update(row.keys())
    all_keys.add('campaign_name')
    
    columnas = sorted(list(all_keys))
    if 'campaign_name' in columnas:
        columnas.remove('campaign_name')
        columnas.append('campaign_name')
    
    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "AMD Detalle"
    
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws.append(columnas)
    for i in range(1, len(columnas) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
    
    for row in rows:
        ws.append([str(row.get(col, '')) for col in columnas])
    
    for i in range(1, len(columnas) + 1):
        letter = get_column_letter(i)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=i).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(45, max(10, max_len + 2))
    
    for r in range(2, ws.max_row + 1):
        for i in range(1, len(columnas) + 1):
            ws.cell(row=r, column=i).alignment = left
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def guardar(contenido, servidor, fecha):
    mes = fecha[:7]
    dir_path = os.path.join(BASE_DIR, mes, fecha, "campanas")
    os.makedirs(dir_path, exist_ok=True)
    ruta = os.path.join(dir_path, f"{servidor}_campaign_all_{fecha}.xlsx")
    with open(ruta, "wb") as f:
        f.write(contenido)
    logger.info(f"   💾 Guardado: {os.path.basename(ruta)}")
    return True

def procesar_fecha(fecha):
    logger.info(f"\n{'='*50}")
    logger.info(f"📅 FECHA: {fecha}")
    logger.info(f"{'='*50}")
    
    for nombre in SERVIDORES:
        servidor = get_server(nombre)
        if not servidor:
            logger.warning(f"⚠️ Servidor {nombre} no configurado")
            continue
        
        logger.info(f"\n📡 Servidor: {nombre}")
        contenido = descargar_reporte(servidor, fecha)
        
        if contenido:
            excel = crear_excel_con_nombres(contenido, servidor, fecha)
            if excel:
                guardar(excel, nombre, fecha)
        else:
            logger.info(f"   ⏭️ Sin datos para {nombre}")
        
        time.sleep(1)

# ========== GENERAR FECHAS ==========
def generar_fechas(inicio, fin):
    start = datetime.strptime(inicio, "%Y-%m-%d")
    end = datetime.strptime(fin, "%Y-%m-%d")
    return [(start + timedelta(days=i)).strftime("%Y-%m-%d") 
            for i in range((end - start).days + 1)]

# ========== EJECUTAR ==========
fechas = generar_fechas(FECHA_INICIO, FECHA_FIN)

logger.info("="*70)
logger.info("🚀 INICIANDO DESCARGA AMD CON NOMBRES DE CAMPAÑA")
logger.info(f"📅 Rango: {FECHA_INICIO} → {FECHA_FIN} ({len(fechas)} días)")
logger.info(f"📁 Ruta base: {BASE_DIR}")
logger.info("="*70)

for fecha in fechas:
    procesar_fecha(fecha)

logger.info("\n" + "="*70)
logger.info("🏁 PROCESO COMPLETADO")
logger.info("="*70)