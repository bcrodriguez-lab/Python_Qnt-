#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Módulo de descarga automática de reportes CDR (llamadas)
Usa config.py para todas las configuraciones.
"""

import os
import time
import logging
import requests
import json
import schedule
import threading
import io
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Importar configuración central
from config import (
    BASE_DIR,
    HORARIOS_EJECUCION,
    DESCARGAR_CDR,
    obtener_fechas_descarga,
    obtener_servidores,
    obtener_token,
    obtener_url_base,
    CONFIG_JSON
)

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger(__name__)

# Mapeo de cortes (fijo)
CORTE_A_SERVIDOR = {
    1: "operacion-interna",
    2: "qnt_juridico_blaster",
    3: "qnt_cobro_blaster",
    4: "Qnt_RBK_blaster",
    5: "Qnt_recaudo_blaster",
    6: "qnt_digital",
}

_scheduler_running = False
_scheduler_thread = None
_descarga_lock = threading.Lock()

# ========== FUNCIONES ==========

def _to_wolkvox_ts(s: str, is_end: bool = False) -> str:
    if not s:
        return ''
    try:
        if 'T' in s or len(s) > 10:
            dt = datetime.fromisoformat(s)
        else:
            d = datetime.strptime(s, '%Y-%m-%d')
            dt = datetime(d.year, d.month, d.day, 23, 59, 59) if is_end else datetime(d.year, d.month, d.day, 0, 0, 0)
        return dt.strftime('%Y%m%d%H%M%S')
    except Exception:
        return s

def obtener_servidores_activos() -> List[Dict]:
    servidores = obtener_servidores()
    return [s for s in servidores if s.get('name') in CORTE_A_SERVIDOR.values()]

def descargar_reporte(servidor_nombre: str, fecha: str) -> Optional[Dict]:
    srv = next((s for s in obtener_servidores() if s.get('name') == servidor_nombre), None)
    if not srv:
        logger.error(f"❌ Servidor {servidor_nombre} no encontrado")
        return None

    url_base = srv.get('url', '').strip().rstrip('/')
    if url_base.lower().startswith('http'):
        base_url = url_base
    else:
        base_url = f"https://wv{url_base}.wolkvox.com"

    token = srv.get('token') or CONFIG_JSON.get('wolkvox-token')
    if not token:
        logger.error(f"❌ No hay token para {servidor_nombre}")
        return None

    date_ini_ts = _to_wolkvox_ts(fecha, is_end=False)
    date_end_ts = _to_wolkvox_ts(fecha, is_end=True)

    url = f"{base_url}/api/v2/reports_manager.php?api=cdr_1&date_ini={date_ini_ts}&date_end={date_end_ts}"
    headers = {"wolkvox-token": token}

    logger.info(f"    📡 Descargando {servidor_nombre} para {fecha}")
    try:
        response = requests.get(url, headers=headers, timeout=60)
        if response.status_code != 200:
            logger.error(f"    ❌ Error {response.status_code}: {response.text[:200]}")
            return None
        data = response.json()
        if isinstance(data, dict) and data.get('data'):
            return data
        elif isinstance(data, list):
            return {'data': data}
        else:
            logger.error("    ❌ Formato de respuesta no reconocido")
            return None
    except Exception as e:
        logger.error(f"    ❌ Error: {e}")
        return None

def convertir_json_a_xlsx(data: Dict, fecha: str, servidor_nombre: str) -> Optional[bytes]:
    rows = data.get('data', [])
    if not rows:
        logger.warning(f"    ⚠️ Sin datos para {servidor_nombre}")
        return None

    all_keys = set()
    for row in rows:
        all_keys.update(row.keys())
    columnas = sorted(list(all_keys))
    if "servidor" not in columnas:
        columnas.append("servidor")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.append(columnas)
    for col_idx in range(1, len(columnas) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in rows:
        fila = [str(row.get(col, '')) for col in columnas[:-1]]
        fila.append(servidor_nombre)
        ws.append(fila)

    for col_idx in range(1, len(columnas) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        width = min(45, max(10, max_len + 2))
        ws.column_dimensions[letter].width = width

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columnas) + 1):
            ws.cell(row=r, column=c).alignment = left

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def guardar_reporte_xlsx(contenido: bytes, servidor_nombre: str, fecha: str) -> bool:
    mes = fecha[:7]
    fecha_dir = os.path.join(BASE_DIR, mes, fecha)
    os.makedirs(fecha_dir, exist_ok=True)
    nombre = f"{servidor_nombre}_{fecha}.xlsx"
    ruta = os.path.join(fecha_dir, nombre)
    with open(ruta, "wb") as f:
        f.write(contenido)
    logger.info(f"      ✅ Guardado: {nombre}")
    return True

def descargar_todos_los_reportes(fecha: str = None):
    if fecha is None:
        fecha = datetime.now().strftime("%Y-%m-%d")

    logger.info(f"\n{'='*60}")
    logger.info(f"📥 INICIANDO DESCARGA DE REPORTES (CDR)")
    logger.info(f"📅 Fecha: {fecha}")
    logger.info(f"📁 Ruta: {os.path.join(BASE_DIR, fecha[:7], fecha)}")
    logger.info(f"{'='*60}")

    if not os.path.exists(BASE_DIR):
        logger.error(f"❌ La ruta base {BASE_DIR} no existe")
        return

    servidores = obtener_servidores_activos()
    if not servidores:
        logger.warning("⚠️ No hay servidores activos")
        return

    total_exitosos = 0
    total_fallidos = 0

    for idx, servidor in enumerate(servidores):
        nombre = servidor.get('name')
        logger.info(f"\n📂 Procesando servidor: {nombre}")
        data = descargar_reporte(nombre, fecha)
        if data is None:
            total_fallidos += 1
            continue

        xlsx = convertir_json_a_xlsx(data, fecha, nombre)
        if xlsx is None:
            total_fallidos += 1
            continue

        if guardar_reporte_xlsx(xlsx, nombre, fecha):
            total_exitosos += 1
        else:
            total_fallidos += 1

        time.sleep(3)

    logger.info(f"\n📊 RESUMEN: ✅ {total_exitosos} exitosos, ❌ {total_fallidos} fallidos")

def descargar_segun_configuracion():
    """Descarga según la configuración del archivo config.py"""
    fechas = obtener_fechas_descarga()
    for fecha in fechas:
        descargar_todos_los_reportes(fecha)

def ejecutar_descarga_programada():
    if not DESCARGAR_CDR:
        logger.info("⏭️ Descargas CDR desactivadas en config.py")
        return
    try:
        descargar_segun_configuracion()
    except Exception as e:
        logger.error(f"❌ Error en descarga programada CDR: {e}")

def iniciar_scheduler():
    global _scheduler_running, _scheduler_thread
    if _scheduler_running:
        logger.info("⚠️ El scheduler CDR ya está en ejecución")
        return False

    if not os.path.exists(BASE_DIR):
        logger.error(f"❌ La ruta base {BASE_DIR} no existe")
        return False

    servidores = obtener_servidores_activos()
    if not servidores:
        logger.warning("⚠️ No hay servidores activos")
        return False

    def run_scheduler():
        global _scheduler_running
        _scheduler_running = True
        logger.info("🚀 INICIANDO SISTEMA DE DESCARGAS PROGRAMADAS (CDR)")
        logger.info(f"📋 Modo: {MODO_DESCARGA}")
        fechas = obtener_fechas_descarga()
        logger.info(f"📅 Fechas a descargar: {fechas}")
        for horario in HORARIOS_EJECUCION:
            schedule.every().day.at(horario).do(ejecutar_descarga_programada)
            logger.info(f"  📅 Descarga programada a las {horario}")
        while _scheduler_running:
            schedule.run_pending()
            time.sleep(30)
        logger.info("🛑 Scheduler CDR detenido")

    _scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    _scheduler_thread.start()
    logger.info("✅ Scheduler CDR iniciado en segundo plano")
    return True

def detener_scheduler():
    global _scheduler_running
    if _scheduler_running:
        _scheduler_running = False
        return True
    return False

def estado_scheduler():
    return {
        "running": _scheduler_running,
        "horarios": HORARIOS_EJECUCION,
        "servidores": len(obtener_servidores_activos()),
        "base_dir": BASE_DIR,
        "modo": MODO_DESCARGA,
        "fechas": obtener_fechas_descarga()
    }

def init_auto_download():
    return iniciar_scheduler()

if __name__ == "__main__":
    init_auto_download()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        detener_scheduler()
        logger.info("Script CDR finalizado")