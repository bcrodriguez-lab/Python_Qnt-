#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script para descargar reportes AMD (campaign_7) con nombres de campaña.
USANDO EL ENDPOINT ESPECÍFICO PARA OBTENER NOMBRES (funcionó en las pruebas)
"""

import os
import json
import requests
import io
import re
import time
import schedule
import threading
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Importar configuración central
from config import (
    BASE_DIR,
    HORARIOS_EJECUCION,
    DESCARGAR_AMD,
    MODO_DESCARGA,
    obtener_fechas_descarga,
    obtener_servidores,
    CONFIG_JSON
)

# Configuración de logging
import logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler("download_campaign_detail.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

SERVIDORES = [
    "operacion-interna",
    "qnt_juridico_blaster",
    "qnt_cobro_blaster",
    "Qnt_RBK_blaster",
    "Qnt_recaudo_blaster",
    "qnt_digital"
]

_scheduler_running_amd = False
_scheduler_thread_amd = None
_descarga_lock_amd = threading.Lock()

# ========== FUNCIONES ==========

def get_server(name):
    for s in CONFIG_JSON.get('servers', []):
        if s.get('name') == name:
            url = s.get('url', '').strip().rstrip('/')
            token = s.get('token') or CONFIG_JSON.get('wolkvox-token')
            if url and token:
                return {'name': name, 'url': url, 'token': token}
    return None

def obtener_nombre_campana_especifica(servidor, campaign_id):
    """Obtiene el nombre de una campaña específica usando el endpoint real_time."""
    try:
        camp_id_clean = re.sub(r'[^0-9]', '', str(campaign_id))
        if not camp_id_clean:
            return None

        url = f"{servidor['url']}/api/v2/real_time.php?api=campaigns&campaign_id={campaign_id}"
        headers = {"wolkvox-token": servidor['token']}

        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            campaigns = data.get('data', [])
            
            # Buscar la campaña que coincide con el ID
            for camp in campaigns:
                camp_str = camp.get('campaign', '')
                if camp_str and camp_str.startswith(camp_id_clean):
                    parts = camp_str.split(' - ', 1)
                    nombre = parts[1] if len(parts) > 1 else camp_str
                    # Limpiar nombre (quitar "Strategy")
                    nombre_limpio = re.sub(r'\s*-\s*Strategy\s*\([^)]*\)\s*$', '', nombre)
                    nombre_limpio = re.sub(r'\s*-\s*[a-zA-Z0-9_]+\s*\([^)]*\)\s*$', '', nombre_limpio)
                    return nombre_limpio if nombre_limpio else nombre
            
            # Fallback: usar el primer resultado si no hay coincidencia exacta
            if campaigns:
                camp_str = campaigns[0].get('campaign', '')
                if camp_str:
                    parts = camp_str.split(' - ', 1)
                    return parts[1] if len(parts) > 1 else camp_str
        return None
    except Exception as e:
        logger.debug(f"Error obteniendo nombre para {campaign_id}: {e}")
        return None

def to_ts(fecha, is_end=False):
    try:
        d = datetime.strptime(fecha, '%Y-%m-%d')
        dt = datetime(d.year, d.month, d.day, 23, 59, 59) if is_end else datetime(d.year, d.month, d.day, 0, 0, 0)
        return dt.strftime('%Y%m%d%H%M%S')
    except:
        return fecha

def descargar_reporte(servidor, fecha):
    url = f"{servidor['url']}/api/v2/reports_manager.php?api=campaign_7&campaign_id=ALL&date_ini={to_ts(fecha)}&date_end={to_ts(fecha, True)}"
    try:
        resp = requests.get(url, headers={"wolkvox-token": servidor['token']}, timeout=60)
        if resp.status_code == 200:
            data = resp.json()
            if data.get('data'):
                return resp.content
        return None
    except Exception as e:
        logger.error(f"Error descargando: {e}")
        return None

def crear_excel_con_nombres(data_bytes, servidor, fecha):
    data = json.loads(data_bytes)
    rows = data.get('data', [])
    if not rows:
        return None

    # === 1. Extraer IDs únicos ===
    ids_unicos = set()
    for row in rows:
        camp_id = str(row.get('campaign_id', '')).strip()
        if camp_id:
            camp_id_clean = re.sub(r'[^0-9]', '', camp_id)
            if camp_id_clean:
                ids_unicos.add(camp_id_clean)

    logger.info(f"   📋 {len(ids_unicos)} IDs de campaña únicos encontrados")

    # === 2. Obtener nombres de la API ===
    mapeo = {}
    for camp_id in ids_unicos:
        nombre = obtener_nombre_campana_especifica(servidor, camp_id)
        if nombre:
            mapeo[camp_id] = nombre
            logger.info(f"      ✅ {camp_id} → {nombre}")
        else:
            logger.warning(f"      ⚠️ {camp_id} → Sin nombre")
        time.sleep(0.5)

    logger.info(f"   📋 {len(mapeo)} nombres obtenidos de {len(ids_unicos)} IDs")

    # === 3. Aplicar nombres a los datos ===
    for row in rows:
        camp_id_original = str(row.get('campaign_id', '')).strip()
        camp_id_clean = re.sub(r'[^0-9]', '', camp_id_original)
        
        if camp_id_clean and camp_id_clean in mapeo:
            nombre = mapeo[camp_id_clean]
            row['campaign_id'] = f"{camp_id_clean} - {nombre}"
            row['campaign_name'] = nombre
        else:
            row['campaign_name'] = ''

    # === 4. Recolectar columnas ===
    all_keys = set()
    for row in rows:
        all_keys.update(row.keys())
    all_keys.add('campaign_name')

    columnas = sorted(list(all_keys))
    if 'campaign_name' in columnas:
        columnas.remove('campaign_name')
        columnas.append('campaign_name')

    # === 5. Crear Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "AMD Detalle"

    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.append(columnas)
    for i in range(1, len(columnas) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = fill
        cell.font = font
        cell.alignment = center

    for row in rows:
        ws.append([str(row.get(col, '')) for col in columnas])

    for i in range(1, len(columnas) + 1):
        letter = get_column_letter(i)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=i).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(45, max(10, max_len + 2))

    for r in range(2, ws.max_row + 1):
        for i in range(1, len(columnas) + 1):
            ws.cell(row=r, column=i).alignment = left

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def guardar(contenido, servidor, fecha):
    mes = fecha[:7]
    dir_path = os.path.join(BASE_DIR, mes, fecha, "campanas")
    os.makedirs(dir_path, exist_ok=True)
    ruta = os.path.join(dir_path, f"{servidor}_campaign_ALL_{fecha}.xlsx")
    with open(ruta, "wb") as f:
        f.write(contenido)
    logger.info(f"   💾 Guardado: {os.path.basename(ruta)}")
    return True

def descargar_todos_los_reportes_amd(fecha: str = None):
    with _descarga_lock_amd:
        if fecha is None:
            fecha = datetime.now().strftime("%Y-%m-%d")

        logger.info(f"\n{'='*60}")
        logger.info(f"📥 INICIANDO DESCARGA DE REPORTES AMD")
        logger.info(f"📅 Fecha: {fecha}")
        logger.info(f"{'='*60}")

        if not os.path.exists(BASE_DIR):
            logger.error(f"❌ La ruta base {BASE_DIR} no existe")
            return

        total_exitosos = 0
        total_fallidos = 0

        for nombre in SERVIDORES:
            servidor = get_server(nombre)
            if not servidor:
                logger.warning(f"⚠️ Servidor {nombre} no configurado")
                continue

            logger.info(f"\n📡 Servidor: {nombre}")
            contenido = descargar_reporte(servidor, fecha)
            if contenido is None:
                logger.info(f"   ⏭️ Sin datos para {nombre}")
                total_fallidos += 1
                continue

            excel = crear_excel_con_nombres(contenido, servidor, fecha)
            if excel:
                if guardar(excel, nombre, fecha):
                    total_exitosos += 1
                else:
                    total_fallidos += 1
            else:
                total_fallidos += 1

            time.sleep(1)

        logger.info(f"\n📊 RESUMEN AMD - Fecha {fecha}")
        logger.info(f"  ✅ Éxitos: {total_exitosos}")
        logger.info(f"  ❌ Fallos: {total_fallidos}")

def descargar_segun_configuracion_amd():
    if not DESCARGAR_AMD:
        logger.info("⏭️ Descargas AMD desactivadas en config.py")
        return
    fechas = obtener_fechas_descarga()
    for fecha in fechas:
        descargar_todos_los_reportes_amd(fecha)

def ejecutar_descarga_programada_amd():
    try:
        descargar_segun_configuracion_amd()
    except Exception as e:
        logger.error(f"❌ Error en descarga programada AMD: {e}")

def iniciar_scheduler_amd():
    global _scheduler_running_amd, _scheduler_thread_amd
    if _scheduler_running_amd:
        logger.info("⚠️ El scheduler AMD ya está en ejecución")
        return False

    if not os.path.exists(BASE_DIR):
        logger.error(f"❌ La ruta base {BASE_DIR} no existe")
        return False

    logger.info("="*60)
    logger.info("📋 CONFIGURACIÓN DE DESCARGA AMD")
    logger.info("="*60)
    logger.info(f"  📅 Modo: {MODO_DESCARGA}")
    fechas = obtener_fechas_descarga()
    logger.info(f"  📅 Fechas a descargar: {fechas}")
    logger.info("="*60)

    def run_scheduler_amd():
        global _scheduler_running_amd
        _scheduler_running_amd = True

        logger.info("="*60)
        logger.info("🚀 INICIANDO SISTEMA DE DESCARGAS PROGRAMADAS (AMD)")
        logger.info("="*60)
        logger.info(f"📁 Ruta base: {BASE_DIR}")
        logger.info(f"📋 Servidores: {SERVIDORES}")

        for horario in HORARIOS_EJECUCION:
            schedule.every().day.at(horario).do(ejecutar_descarga_programada_amd)
            logger.info(f"  📅 Descarga AMD programada a las {horario}")

        logger.info(f"\n🔄 Scheduler AMD iniciado. Próximas ejecuciones:")
        for horario in HORARIOS_EJECUCION:
            logger.info(f"  ⏰ {horario}")

        while _scheduler_running_amd:
            schedule.run_pending()
            time.sleep(30)

        logger.info("🛑 Scheduler AMD detenido")

    _scheduler_thread_amd = threading.Thread(target=run_scheduler_amd, daemon=True)
    _scheduler_thread_amd.start()
    logger.info("✅ Scheduler AMD iniciado en segundo plano")
    return True

def detener_scheduler_amd():
    global _scheduler_running_amd
    if _scheduler_running_amd:
        _scheduler_running_amd = False
        logger.info("🛑 Deteniendo scheduler AMD...")
        return True
    return False

def estado_scheduler_amd() -> Dict:
    return {
        "running": _scheduler_running_amd,
        "horarios": HORARIOS_EJECUCION,
        "servidores": SERVIDORES,
        "base_dir": BASE_DIR,
        "modo_descarga": MODO_DESCARGA,
        "fechas": obtener_fechas_descarga()
    }

def init_auto_download_amd():
    return iniciar_scheduler_amd()

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    init_auto_download_amd()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        detener_scheduler_amd()
        logger.info("Script AMD finalizado")